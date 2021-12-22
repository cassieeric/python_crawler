# -*- coding: utf-8 -*-
# @Author: hebe
# @Date:   2020-04-18 18:31:03
# @Last Modified by:   hebe
# @Last Modified time: 2020-04-18 19:40:48
import os 
import glob
import openpyxl

def merge_xlsx_files(xlsx_files):
    wb = openpyxl.load_workbook(xlsx_files[0])
    ws = wb.active
    ws.title = "merged result"

    for  filename in xlsx_files[1:]:
        workbook = openpyxl.load_workbook(filename)
        sheet = workbook.active
        for row in sheet.iter_rows(min_row=1):
            values = [cell.value for cell in row]
            ws.append(values)
    return wb

#path is very important here , must true.
def get_all_xlsx_files(path):
    xlsx_files = glob.glob(os.path.join(r'E:\PythonCrawler\python_crawler-master\MergeExcelSheet\testfile\\file\\', '*.xlsx'))
    sorted(xlsx_files, key=str.lower)
    return xlsx_files

def main():
    xlsx_files = get_all_xlsx_files(os.path.expanduser('~lmx'))
    wb = merge_xlsx_files(xlsx_files)
    wb.save('Jayson提供的代码.xlsx')


if __name__ == '__main__':
    main()
    
print("all excel append OK!")
