# -*- coding: utf-8 -*-
from openpyxl import load_workbook, Workbook
import glob

path = "C:\\Users\\pdcfi\\Desktop\\投稿\\崔艳飞\\投稿4-Python实现多Excel多Sheet合并\\excel\\"
new_workbook = Workbook()
new_sheet = new_workbook.active

# 用flag变量明确新表是否已经添加了表头，只要添加过一次就无须重复再添加
flag = 0

for file in glob.glob(path + '/*.xlsx'):
    workbook = load_workbook(file)
    sheet = workbook.active

    coloum_A = sheet['A']
    row_lst = []
    for cell in coloum_A:
        if cell:
            print(cell.row)
            row_lst.append(cell.row)

    if not flag:
        header = sheet[1]
        header_lst = []
        for cell in header:
            header_lst.append(cell.value)
        new_sheet.append(header_lst)
        flag = 1

    for row in row_lst:
        data_lst = []
        for cell in sheet[row]:
            data_lst.append(cell.value)
        new_sheet.append(data_lst)

new_workbook.save(path + '/' + '符合筛选条件的新表.xlsx')